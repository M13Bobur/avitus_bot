import os
import tempfile
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.logging_config import get_logger
from app.repositories.inventory import InventoryRepository

logger = get_logger(__name__)

EXPORT_COLUMNS = ["Филиал", "Наименование", "Кол-во", "Дата"]


def _format_excel_columns(file_path: str) -> None:
  workbook = load_workbook(file_path)
  worksheet = workbook.active

  for col_idx, header in enumerate(EXPORT_COLUMNS, start=1):
    max_length = len(header)
    for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
      for cell in row:
        if cell.value is not None:
          max_length = max(max_length, len(str(cell.value)))

    worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 60)

  workbook.save(file_path)
  workbook.close()


class ExcelExportService:
  def __init__(self, inventory_repo: InventoryRepository) -> None:
    self._inventory_repo = inventory_repo

  async def generate_supplier_report(self, supplier_id: int) -> tuple[str, int]:
    records = await self._inventory_repo.get_by_supplier(supplier_id)

    data = [
      {
        "Филиал": record.branch.name,
        "Наименование": record.medicine.name,
        "Кол-во": float(record.quantity),
        "Дата": record.report_date.strftime("%d.%m.%Y"),
      }
      for record in records
    ]

    df = pd.DataFrame(data, columns=EXPORT_COLUMNS)

    tmp_dir = tempfile.mkdtemp()
    file_path = os.path.join(
      tmp_dir,
      f"inventar_{supplier_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    df.to_excel(file_path, index=False, engine="openpyxl")
    if data:
      _format_excel_columns(file_path)

    logger.info(
      "excel_export_generated",
      supplier_id=supplier_id,
      rows=len(data),
      file_path=file_path,
    )
    return file_path, len(data)
